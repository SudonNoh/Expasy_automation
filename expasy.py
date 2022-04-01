from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date
import pandas as pd
import os

class SeleniumControl:
    def __init__(self, version):
        self.chrome_service = Service(ChromeDriverManager(version=version).install())
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('windows-size=800,1000')
        self.driver = webdriver.Chrome(service=self.chrome_service, options=self.options)

    def site_enter(self, url):
        self.driver.get(url=url)
        self.driver.implicitly_wait(5)

    def input_seq(self, seq):
        self.search = self.driver.find_element(By.XPATH, '//*[@id="sib_body"]/form/textarea')
        self.search.clear()
        self.search.send_keys(seq)
        self.driver.find_element(By.XPATH, '//*[@id="sib_body"]/form/p[1]/input[2]').submit()

    def get_body(self):
        body = self.driver.find_element(By.XPATH, '//*[@id="sib_body"]')
        return body.text

    def time_sleep(self, sec):
        self.driver.implicitly_wait(sec)

    def site_back(self):
        self.driver.back()

    def site_close(self):
        self.driver.quit()


class ExcelControl:
    
    def excel_read(self, url, sheet_name):
        # excel 불러오기
        excel_data = pd.read_excel(url, sheet_name=sheet_name)
        # 비어있는 행 파일 정리
        data = excel_data.drop(excel_data.index[0:3])
        # 사용할 열을 추출하고 NA값 제거
        data = data[[data.columns[7], data.columns[8]]].dropna(axis=0)
        data1 = data[[data.columns[1]]].dropna(axis=0)
        data2 = data[[data.columns[0]]].dropna(axis=0)
        # 데이터 안에 있는 값들을 하나씩 추출해 리스트로 저장
        seq = [j for i in data1.values.tolist() for j in i]
        seq_id = [j for i in data2.values.tolist() for j in i]
        # sequence list로 추출
        return seq, seq_id
    
    def save_excel_file(self, data_list, seq_id, url, sheet_name):
        
        # MW, PI, ABS
        data_frame_list = []
        for i, j in zip(data_list, seq_id):
            # MW, PI, ABS
            data = []
            data.append(i)
            data.append(j)
            data.append(round(float(self.string_slice(i, 'Molecular weight:'))/1000, 2))
            data.append(self.string_slice(i, 'Theoretical pI:'))
            data.append(self.string_slice(i, 'Abs 0.1% (=1 g/l)'))
            data_frame_list.append(data)
            
        df = pd.DataFrame(data_frame_list, columns=['RESULT', 'ID', 'MW', 'PI', 'Abs'])
        new_url = self.make_new_url(url)
        writer = pd.ExcelWriter(new_url, engine='openpyxl')
        df.to_excel(writer, sheet_name="ExpasyProtParam", index=False)
        writer.save()

    # 결과값을 String으로 전체를 보여주는 함수
    def make_excel_file(self, data_list, url, sheet_name):
        
        # data list를 불러와서 하나의 string으로 추출
        string = ''
        for i in data_list:
            string += i+'\n\n'
            
        data_list = string.split('\n')
        df = pd.DataFrame(data_list)
        
        wb = Workbook()
        # wb.create_sheet(title=sheet_name)
        ws = wb['Sheet']
        ws.title = sheet_name
        
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
            
        new_url = self.make_new_url(url)
        
        wb.save(filename=new_url)
        
    # 결과값으로 MW, PI, ABS 를 보여주는 함수
    def make_excel_file_2(self, data_list, url, sheet_name, seq_id):
        
        # MW, PI, ABS
        data_frame_list = []
        for i, j in zip(data_list, seq_id):
            data = []
            data.append(j)
            data.append(round(float(self.string_slice(i, 'Molecular weight:'))/1000, 2))
            data.append(self.string_slice(i, 'Theoretical pI:'))
            data.append(self.string_slice(i, 'Abs 0.1% (=1 g/l)'))
            data_frame_list.append(data)
            
        # data list를 불러와서 하나의 string으로 추출
        string = ''
        for i in data_list:
            string += i+'\n\n'
            
        datas = string.split('\n')
        
        df = pd.DataFrame(data_frame_list, columns=['ID', 'MW', 'PI', 'Abs'])
        df2 = pd.DataFrame(datas)
        
        new_url = self.make_new_url(url)
        
        writer = pd.ExcelWriter(new_url, engine='openpyxl')
        df.to_excel(writer, sheet_name="MW", index=False)
        df2.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
        
    def string_slice(self, data, string):
        try:
            str_len = len(string)
            start_num = data.find(string)+str_len
            if 'Abs' in string:
                export_string = data[start_num:data.find(',', start_num)]
                if 'Estimatedhalf-life:' in export_string:
                    export_string = data[start_num:data.find('Estimatedhalf-life:', start_num)]
            else:
                export_string = data[start_num:data.find('\n', start_num)]
        except:
            return ' '
        return export_string.replace(' ', '')
        
    def make_new_url(self, url):
        
        today = date.today().strftime('%y%m%d')
        
        folder = url[:url.rfind("/")]
        file_list = os.listdir(folder)
        
        count = 1
        if url[url.rfind("/")+1:] in file_list:
            for i in file_list:
                if today in i:
                    count += 1
        else:
            count = 1
        new_url = url[:url.rfind('/')+1]+'ExpasyProtParam' + '_' + today + '(' + str(count) + ')' + '.xlsx'
        
        return new_url